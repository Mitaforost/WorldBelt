import requests
from bs4 import BeautifulSoup
from urllib.parse import urljoin
from openpyxl import Workbook
import re
import time


def sanitize_sheet_title(title):
    """Удаляем недопустимые символы и обрезаем до 31 символа (ограничение Excel)."""
    title = re.sub(r'[:\\/?*\[\]]', '', title)
    return title[:31]


def extract_text_and_structure(url, wb, menu_sheet, visited=None):
    if visited is None:
        visited = set()

    if url in visited:
        return
    visited.add(url)

    try:
        response = requests.get(url, timeout=10)
    except Exception as e:
        print(f"Ошибка при загрузке {url}: {e}")
        return

    if response.status_code != 200:
        print(f"Ошибка при загрузке {url} — код {response.status_code}")
        return

    soup = BeautifulSoup(response.text, 'html.parser')

    # Название категории
    h1_tag = soup.find('h1')
    category_name = h1_tag.text.strip() if h1_tag else "Без названия"
    sheet_title = sanitize_sheet_title(category_name)

    print(f"Обрабатывается категория: {category_name}")

    # Создание листа категории
    if sheet_title not in wb.sheetnames:
        category_sheet = wb.create_sheet(title=sheet_title)
        category_sheet.append(["Ключевые слова", "Название товаров", "Ссылки на товары"])
    else:
        category_sheet = wb[sheet_title]

    # Добавление ссылки в меню
    excel_link = f'=HYPERLINK("#\'{sheet_title}\'!A1", "{category_name}")'
    menu_sheet.append([excel_link, url])

    # Поиск карточек товаров
    product_cards = soup.select('.product-card')
    for card in product_cards:
        name_tag = card.select_one('.card-title')
        link_tag = card.get('href') or card.find('a')
        keyword = name_tag.text.strip() if name_tag else "Без названия"
        link = urljoin(url, link_tag.get('href')) if link_tag and link_tag.get('href') else url

        category_sheet.append([keyword, keyword, link])

    # Поиск подкатегорий
    subcategories = soup.select('a.btn.btn-sm.btn-outline-secondary')
    for sub in subcategories:
        sub_url = urljoin(url, sub.get('href'))
        if sub_url not in visited:
            time.sleep(1)
            extract_text_and_structure(sub_url, wb, menu_sheet, visited)


if __name__ == "__main__":
    start_url = input("Введите URL сайта: ").strip()

    wb = Workbook()
    menu_sheet = wb.active
    menu_sheet.title = "Меню"
    menu_sheet.append(["Категория", "URL"])

    extract_text_and_structure(start_url, wb, menu_sheet)

    filename = "site_structure.xlsx"
    wb.save(filename)
    print(f"✅ Структура сайта сохранена в {filename}")
