import requests
from bs4 import BeautifulSoup
import matplotlib.pyplot as plt
from collections import Counter
import openpyxl
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd
import nltk
from nltk.corpus import stopwords
import re
import os

nltk.download('stopwords')

def fetch_html(url):
    headers = {"User-Agent": "Mozilla/5.0"}
    response = requests.get(url, headers=headers)
    return response.text

def parse_meta(soup):
    meta_info = {
        "title": soup.title.string if soup.title else "",
        "description": "",
        "keywords": ""
    }
    for tag in soup.find_all("meta"):
        if tag.get("name") == "description":
            meta_info["description"] = tag.get("content", "")
        elif tag.get("name") == "keywords":
            meta_info["keywords"] = tag.get("content", "")
    return meta_info

def parse_headers(soup):
    headers = []
    for i in range(1, 7):
        for tag in soup.find_all(f'h{i}'):
            headers.append((f'H{i}', tag.get_text(strip=True)))
    return headers

def extract_text(soup):
    for script in soup(["script", "style"]):
        script.decompose()
    return soup.get_text()

def word_density(text):
    text = re.sub(r'[^\w\s]', '', text.lower())
    words = text.split()
    filtered_words = [w for w in words if w not in stopwords.words("russian") and w not in stopwords.words("english")]
    return Counter(filtered_words)

def save_to_excel(meta, headers, word_counts, graph_path, filename="seo_analysis.xlsx"):
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "Метатеги"
    ws1.append(["Тег", "Содержимое"])
    for key, value in meta.items():
        ws1.append([key, value])

    ws2 = wb.create_sheet("Заголовки")
    ws2.append(["Тип", "Текст"])
    for h_type, text in headers:
        ws2.append([h_type, text])

    ws3 = wb.create_sheet("Плотность слов")
    df = pd.DataFrame(word_counts.most_common(20), columns=["Слово", "Частота"])
    for r in dataframe_to_rows(df, index=False, header=True):
        ws3.append(r)

    img = ExcelImage(graph_path)
    ws3.add_image(img, "D2")

    wb.save(filename)

def plot_word_density(counter, path='word_density.png'):
    top_words = counter.most_common(20)
    words, freqs = zip(*top_words)
    plt.figure(figsize=(12, 6))
    plt.bar(words, freqs, color='skyblue')
    plt.title("Плотность ключевых слов")
    plt.xticks(rotation=45, ha="right")
    plt.tight_layout()
    plt.savefig(path)
    plt.close()

def analyze_site(url):
    html = fetch_html(url)
    soup = BeautifulSoup(html, "html.parser")
    meta = parse_meta(soup)
    headers = parse_headers(soup)
    text = extract_text(soup)
    word_counts = word_density(text)
    plot_path = "word_density.png"
    plot_word_density(word_counts, plot_path)
    save_to_excel(meta, headers, word_counts, plot_path)

# 🔧 Пример использования
if __name__ == "__main__":
    url = input("Введите URL сайта: ")
    analyze_site(url)
    print("✅ Анализ завершён. Данные сохранены в seo_analysis.xlsx")
